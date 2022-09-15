import pandas as pd
import openpyxl
import numpy as np
import logging
import re
import sys
from openpyxl.styles.borders import Border, Side
# 文字の位置を扱うにインポートする必要がある
from openpyxl.styles import Alignment
# 文字の太さを扱うにインポートする必要がある
from openpyxl.styles import Font
import os

import tkinter
from tkinter import messagebox

# 例外処理
# Excelファイルが開いている時(開いていると処理できずプログラム落ちる)
# 読み込んだExcelファイルのヘッダーに、機番、作業区分名１、受付日時、車種の何れかが無い(4項目の値で連番付与ロジックを組み立て)
# setting内のexcelファイルが無い　3コード

# その他エラー処理
# Excelがフォルダ内に無い
# ExcelのA1セルに 保全履歴一覧 入力が無い(処理対象のExcelファイルか確認)
# 抽出項目が自動付与したExcelに存在しない

# メインの連番自動付与処理---------------------------------------------
# 引き数ecxel_pathには、ファイル名.xlseが入る

def serial_number_assignment(ecxel_path):
    # 編集するExcelファイルのパスを取得する
    input_original_excel_file_path = './'+ecxel_path

    # A1のみ作成する
    a1 = [["保全履歴一覧", ]]
    a1df = pd.DataFrame(a1, columns=["a"])

# 編集するExcelファイルの左上(A1セル)に保全履歴一覧と入力されていなかったら処理をしない
    # 例外処理 フォルダ内のExcelファイルが開いている時の処理
    try:
        # 変数するExcelファイルをheader = 0 で読み込む
        input_original_excel_file_all = pd.read_excel(
            input_original_excel_file_path, header=0)
    except PermissionError:
        # メインウィンドウの作成
        root = tkinter.Tk()
        # 最前面に表示
        root.attributes('-topmost', True)
        # メインウィンドウを非表示
        root.withdraw()
        # ダイアログの表示
        messagebox.showerror(
            'Excelファイル読み込みエラー', 'Excelファイルが開いている為、処理を実行できません！\nExcelファイルを閉じてから再度実行して下さい。')
        # プログラム強制終了
        sys.exit(1)

    # print("A1セル！ ： " + input_original_excel_file_all.columns[0])
    # 編集するExcelファイルの左上(A1セル)が保全履歴一覧だったらTrue
    if input_original_excel_file_all.columns[0] == "保全履歴一覧":

        # ヘッダーを指定、保全履歴一覧は読み込まない
        input_original_excel_file = pd.read_excel(
            input_original_excel_file_path, header=1)

        try:
            # 必要な列を抽出
            select_original_excel_file_all = input_original_excel_file.loc[:, [
                "機番", "作業区分名１", "受付日時", "車種"]]

        except KeyError:
            # メインウィンドウの作成
            root = tkinter.Tk()
            # 最前面に表示
            root.attributes('-topmost', True)
            # メインウィンドウを非表示
            root.withdraw()
            # ダイアログの表示
            messagebox.showerror(
                'ヘッダーエラー', '読み込んだExcelファイルに、\n機番、作業区分名１、受付日時、車種\nの項目が1つ以上存在しません！\n出力されるExcelファイルの設定変更が無いか確認して下さい。\n設定変更ない場合作業区分名１の数値が半角の可能性があります。\n半角の場合は全角になるように出力側の設定変更をお願いします。')
            # プログラム強制終了
            sys.exit(1)

        # 必要な列を抽出
        select_original_excel_file = input_original_excel_file.loc[:, [
            "機番", "作業区分名１", "受付日時"]]

        # regkiban.xlsxファイルを開く
        input_setting_excel_file_path = 'setting\\regkiban.xlsx'

        try:
            # regkiban.xlsxを読みこむ
            input_setting_excel_file = pd.read_excel(
                input_setting_excel_file_path, sheet_name=None, header=0)
        except FileNotFoundError:
            # メインウィンドウの作成
            root = tkinter.Tk()
            # 最前面に表示
            root.attributes('-topmost', True)
            # メインウィンドウを非表示
            root.withdraw()
            # ダイアログの表示
            messagebox.showerror(
                '設定ファイル読み込みエラー', 'regkiban.xlsxがsettingフォルダ内にありません。\nregkiban.xlsxが存在するか、\nもしくはsettingフォルダが連番自動付与させるexcelファイルと\n同一フォルダ内にあるか確認して下さい。')
            # プログラム強制終了
            sys.exit(1)
        except PermissionError:
            # メインウィンドウの作成
            root = tkinter.Tk()
            # 最前面に表示
            root.attributes('-topmost', True)
            # メインウィンドウを非表示
            root.withdraw()
            # ダイアログの表示
            messagebox.showerror(
                'Excelファイル読み込みエラー', 'Excelファイルが開いている為、処理を実行できません！\nExcelファイルを閉じてから再度実行して下さい。')
            # プログラム強制終了
            sys.exit(1)

        # シート名リストで取得 ['860K日支', 'ステータ', 'Rrモータ', 'PA10', 'PB10']
        input_setting_excel_file_sheet_nameslists = list(
            input_setting_excel_file.keys())

        # 追加する項目の列を新規作成　
        add_list = ["該当機番", "日付のみ", "作業", "連番",
                    "作業1桁", "日付+該当機番+作業", "カウント", "1時連番"]
        for element in add_list:
            select_original_excel_file[element] = ""

        # print(select_original_excel_file)
        # 設定ファイルregwork.xlsxのパスを取得
        input_work_excel_file_path = 'setting\\regwork.xlsx'

        try:
            # regwork.xlsxを読みこむ
            input_work_excel_file = pd.read_excel(
                input_work_excel_file_path, header=0, index_col=0)
        except FileNotFoundError:
            # メインウィンドウの作成
            root = tkinter.Tk()
            # 最前面に表示
            root.attributes('-topmost', True)
            # メインウィンドウを非表示
            root.withdraw()
            # ダイアログの表示
            messagebox.showerror(
                '設定ファイル読み込みエラー', 'regwork.xlsxがsettingフォルダ内にありません。\nregwork.xlsxが存在するか、\nもしくはsettingフォルダが連番自動付与させるexcelファイルと\n同一フォルダ内にあるか確認して下さい。')
            # プログラム強制終了
            sys.exit(1)
        except PermissionError:
            # メインウィンドウの作成
            root = tkinter.Tk()
            # 最前面に表示
            root.attributes('-topmost', True)
            # メインウィンドウを非表示
            root.withdraw()
            # ダイアログの表示
            messagebox.showerror(
                'Excelファイル読み込みエラー', 'Excelファイルが開いている為、処理を実行できません！\nExcelファイルを閉じてから再度実行して下さい。')
            # プログラム強制終了
            sys.exit(1)

        # 作業区分データフレームをdict化
        input_work_excel_file_dicts = input_work_excel_file.to_dict()
        # 作業区分データフレームをlist化
        input_work_excel_file_lists = list(input_work_excel_file["番号"].keys())
        # print("作業区分dict" + input_work_excel_file_dicts)
        # print("作業区分list" + str(input_work_excel_file_lists))

        # 機番列の長さを取得
        select_original_excel_file_machine_numbers_len = len(
            select_original_excel_file["機番"])
        # 機番列の長さだけ繰り返す
        for i in range(select_original_excel_file_machine_numbers_len):
            # 機番列のセルの値を一つずつ取り出す
            select_original_excel_file["機番"][i]

            for element1 in input_setting_excel_file_sheet_nameslists:
                # シートのカラムをリスト化
                input_setting_excel_file_column_lists = list(
                    input_setting_excel_file[element1].columns)
                for element2 in input_setting_excel_file_column_lists:
                    for ii in range(input_setting_excel_file[element1][element2].count()):
                        if ii >= 1:
                            # 比較テストプリント
                            # print(select_original_excel_file["機番"][i]+":"+input_setting_excel_file[element1][element2][ii])
                            # 編集するExcelの機番とregkiban.xlsxが同じだったら
                            if str(select_original_excel_file["機番"][i]) == input_setting_excel_file[element1][element2][ii]:
                                # 比較テストプリント
                                # print(str([i]) + "は" + select_original_excel_file["機番"][i] + ":" + element1 + "は" + input_setting_excel_file[element1][element2][ii])

                                # 該当機番列に付与番号を追加
                                select_original_excel_file["該当機番"][i] = str(
                                    input_setting_excel_file[element1][element2][0])

                                # 該当日付列に日付のみを追加
                                select_original_excel_file_day = select_original_excel_file["受付日時"][i]
                                select_original_excel_file["日付のみ"][i] = str(
                                    select_original_excel_file_day[:10])

                                #  該当作業列に作業区分の番号を追加
                                for element3 in input_work_excel_file_lists:
                                    if str(element3) == select_original_excel_file["作業区分名１"][i]:
                                        select_original_excel_file["作業"][i] = input_work_excel_file_dicts["番号"][element3]

                                if len(str(select_original_excel_file["作業"][i])) == 1:
                                    select_original_excel_file["作業1桁"][i] = 0
                                else:
                                    select_original_excel_file["作業1桁"][i] = "xx"

                                # もし作業セルが空じゃなかったら
                                if select_original_excel_file["作業"][i] != "":

                                    if select_original_excel_file["作業1桁"][i] == 0:

                                        # 日付+該当機番+作業に値追加
                                        select_original_excel_file["日付+該当機番+作業"][i] = select_original_excel_file["日付のみ"][i] + "/" + select_original_excel_file["該当機番"][i] + "-" + str(
                                            select_original_excel_file["作業"][i]) + str(select_original_excel_file["作業1桁"][i])

                                    else:
                                        select_original_excel_file["日付+該当機番+作業"][i] = select_original_excel_file["日付のみ"][i] + "/" + \
                                            select_original_excel_file["該当機番"][i] + "-" + str(
                                                select_original_excel_file["作業"][i])
        # 日付+該当機番+作業行の空欄をNonに変換
        select_original_excel_file["日付+該当機番+作業"].replace(
            "", np.nan, inplace=True)
        # Nonを行を削除
        select_original_excel_file_dropna = select_original_excel_file.dropna()

        # データフレームをカラム名"日付+該当機番+作業","カウント"のみに
        select_original_excel_file_number = select_original_excel_file_dropna.loc[:, [
            "日付+該当機番+作業", "カウント", "1時連番"]]
        # カウントにユニーク連番を付与
        select_original_excel_file_number['カウント'] = select_original_excel_file_number.groupby(
            '日付+該当機番+作業').cumcount()+1

        # テストプリント
        # print(select_original_excel_file_number)

        # データフレームをカラム名"日付+該当機番+作業+1時連"のindexをリスト化
        select_original_excel_file_number_indexlists = list(
            select_original_excel_file_number.index)
        # print(select_original_excel_file_number_indexlists)

        # リスト化したindexを順番に取り出す
        for index in select_original_excel_file_number_indexlists:

            # print(select_original_excel_file_number['カウント'][index])

            select_original_excel_file_number['1時連番'][index] = str(
                select_original_excel_file_number['日付+該当機番+作業'][index]) + str(select_original_excel_file_number['カウント'][index])

            # -以降の数値を取得する
            serial_number = select_original_excel_file_number['1時連番'][index]
            target_string = str('-')
            # 1時連番の文字列から-を見つける
            get_number = serial_number.find(target_string)
            # -以降の文字列を変数に定義
            edit_numbers = serial_number[get_number+1:]
            # print(edit_numbers)

            if len(str(edit_numbers)) == 4:
                # 一文字ずつスライスしてリスト化
                edit_numbers_list = re.split('(.)', edit_numbers)[1::2]
                # edit_numbers_listを一文字ずつ数値型に変換
                edit_numbers_map_int = map(
                    (lambda x: int(x)), edit_numbers_list)
                # 高階関数map使用時、pythonではforかリスト化等しないと中身見れず…面倒…なぜ？
                edit_numbers_list_int = list(edit_numbers_map_int)
                # print(edit_numbers_list_int)
                # index1と2を足す
                edit_number = sum(edit_numbers_list_int[1:3])

                # 再リスト化する空のリスト
                edit_numbers_new_number = ""

                for i in range(len(str(edit_numbers))):
                    if i != 1 and i != 2:
                        edit_numbers_new_number = edit_numbers_new_number + \
                            edit_numbers_list[i]
                    elif i == 1:
                        edit_numbers_new_number = edit_numbers_new_number + \
                            str(edit_number)

                # print("4だった場合 : "+edit_numbers_new_number)

                # 1時連番の文字列から-を見つける
                select_original_excel_file_number_find = select_original_excel_file_number['1時連番'][index].find(
                    target_string)
                # -以降の文字列を変数に定義
                select_original_excel_file_number_new = select_original_excel_file_number[
                    '1時連番'][index][:select_original_excel_file_number_find+1]
                # 1時連番に4桁の数値を3桁にしたものを再度追加
                select_original_excel_file_number['1時連番'][index] = select_original_excel_file_number_new + \
                    edit_numbers_new_number

            # 全て記載されているデータシートのカラム名車種に、1時連番を書き込む
            input_original_excel_file['車種'][index] = select_original_excel_file_number['1時連番'][index]

        # print(select_original_excel_file_all)

        # エクセルファイルに上書き
        input_original_excel_file.to_excel(
            input_original_excel_file_path, index=False)

        # 編集するExcelファイルを開く
        input_original_excel_file_path = './'+ecxel_path

        # 書き込んだExcelを元のフォーマットに戻す
        wb = openpyxl.load_workbook(input_original_excel_file_path)
        # シートの指定
        ws = wb["Sheet1"]
        # 1行目に一行新規挿入
        ws.insert_rows(1)
        # セルA１指定
        rng1 = ws["A1"]
        # A1に"保全履歴一覧"追加
        rng1.value = "保全履歴一覧"

        # セルの背景色設定
        fill = openpyxl.styles.PatternFill(
            patternType='solid', fgColor='FFFF00', bgColor='FFFF00')
        # けい線の設定
        side = Side(style='thin', color='FFFF00')
        # セルの上下左右のけい線指定
        border_no = Border(top=side, bottom=side, left=side, right=side)
        # フォントの太文字解除
        font = Font(bold=False)

        # A2の列からEB2の列に適用
        for rows in ws['A2':'EB2']:
            for cell in rows:
                cell.fill = fill
                cell.border = border_no
                cell.alignment = Alignment(horizontal="left")
                cell.font = font
            # Excelに書き込み
            wb.save(input_original_excel_file_path)

    else:
        # メインウィンドウの作成
        root = tkinter.Tk()
        # 最前面に表示
        root.attributes('-topmost', True)
        # メインウィンドウを非表示
        root.withdraw()
        # ダイアログの表示
        messagebox.showerror('実行エラー', ecxel_path+'\nは連番付与可能なExcelファイルではありません！')
        sys.exit()

# 自動連番付与された項目を抽出する処理-----------------------------------------------------


def serial_number_selection(after_ecxel_path):
    # 抽出する項目が書かれているselect_item.xlsxのパスを取得
    select_item_excel_file_path = 'setting\\select_item.xlsx'

    try:
        # 抽出する項目が書かれているselect_item.xlsxを開く 1行目に注意事項記載の為header=1。2行目から読み込み
        select_excel_file = pd.read_excel(
            select_item_excel_file_path, header=1)
    except FileNotFoundError:
        # メインウィンドウの作成
        root = tkinter.Tk()
        # 最前面に表示
        root.attributes('-topmost', True)
        # メインウィンドウを非表示
        root.withdraw()
        # ダイアログの表示
        messagebox.showerror(
            '設定ファイル読み込みエラー', 'select_item.xlsxがsettingフォルダ内にありません。\nselect_item.xlsxが存在するか、\nもしくはsettingフォルダが連番自動付与させるexcelファイルと\n同一フォルダ内にあるか確認して下さい。')
        # プログラム強制終了
        sys.exit(1)
    except PermissionError:
        # メインウィンドウの作成
        root = tkinter.Tk()
        # 最前面に表示
        root.attributes('-topmost', True)
        # メインウィンドウを非表示
        root.withdraw()
        # ダイアログの表示
        messagebox.showerror(
            'Excelファイル読み込みエラー', 'Excelファイルが開いている為、処理を実行できません！\nExcelファイルを閉じてから再度実行して下さい。')
        # プログラム強制終了
        sys.exit(1)

    # 抽出項目をリスト化する
    select_item_lists = list(select_excel_file['抽出項目'])

    # 抽出項目のリストに、車種　があったらTrue
    select_item_list_bool = '車種' in select_item_lists
    # 抽出項目リストに車種がなかったら追加する
    if select_item_list_bool == False:
        select_item_lists.append('車種')

    # 編集するExcelファイルのパスを取得
    input_original_excel_file_path = './'+after_ecxel_path
    # 編集するExcelファイルを開く
    selection_excel_file = pd.read_excel(
        input_original_excel_file_path, header=1)
    # 編集するExcelファイルのカラム名をリスト化
    selection_excel_file_columnslists = selection_excel_file.columns.tolist()

    # 値を格納する空のリスト作成、初期化
    select_value_lists = []

    # 抽出項目リストの数だけ繰り返す
    for select_item_list in select_item_lists:
        # 抽出項目カウントの初期化
        select_item_list_count = 0
        # カラム名カウントの初期化
        selection_excel_file_columnslist_count = 0
        # 編集するexcelファイルのカラム名リストの数だけ繰り返す
        for selection_excel_file_columnslist in selection_excel_file_columnslists:
            # 抽出項目カウントアップ
            select_item_list_count += 1
            # カラム名カウントアップ
            selection_excel_file_columnslist_count += 1
            # もし、抽出項目と編集するexcelファイルのカラム名がTrueだったら
            if select_item_list == selection_excel_file_columnslist:
                # 値リストに項目追記
                select_value_lists.append(select_item_list)
                # 抽出項目カウントをマイナス
                select_item_list_count -= 1
    #     抽出項目カウントとカラム名カウントが同じだった場合、カラム名に抽出項目が存在しない為、ダイアログ表示。処理は続行。
        if select_item_list_count == selection_excel_file_columnslist_count and str(select_item_list) != "nan":
            # メインウィンドウの作成
            root = tkinter.Tk()
            # 最前面に表示
            root.attributes('-topmost', True)
            # メインウィンドウを非表示
            root.withdraw()
            # ダイアログの表示
            messagebox.showwarning('実行エラー', str(
                select_item_list)+'\nは存在しない項目なのでスキップします。')

    print("削除！" + str(select_value_lists))

    # 転記先の新しいデータフレームを作成し、値リストからカラム追記
    posting_new_df = pd.DataFrame(columns=select_value_lists)

    print(posting_new_df)

    # もし、編集するExcelファイルのB1セルに入力されている値が機番だったら処理を実行
    # if selection_excel_file.columns[0] == "機番":

    # 編集するExcelファイルの左上(A1セル)に保全履歴一覧と入力されていなかったら処理をしない
    # 変数するExcelファイルをheader = 0 で読み込む
    input_original_excel_file_path_all = pd.read_excel(
        input_original_excel_file_path, header=0)
    # print("A1セル！ ： " + input_original_excel_file_all.columns[0])
    # 編集するExcelファイルの左上(A1セル)が保全履歴一覧だったらTrue
    if input_original_excel_file_path_all.columns[0] == "保全履歴一覧":

        print("カラム：　" + selection_excel_file.columns[0])

        print(len(selection_excel_file["機番"]))

        # 編集するExcelファイルの機番の列の数だけ繰り返す selection_excel_file_len = 0～
        for selection_excel_file_len in range(len(selection_excel_file["機番"])):
            # 車種の列のセルを取り出す
            contents_cell = selection_excel_file.loc[selection_excel_file_len, '車種']
            # もし、セルに入力されている値がnanでなかったら
            if str(contents_cell) != "nan":

                # 値を格納する空のリスト作成、初期化
                # select_value_lists = []
                # 値を格納する空のdict作成、初期化
                select_value_dicts = {}

                # 抽出項目リストの数だけ繰り返す
                for select_item_list in select_item_lists:
                    # 編集するexcelファイルのカラム名リストの数だけ繰り返す
                    for selection_excel_file_columnslist in selection_excel_file_columnslists:
                        # もし、抽出項目と編集するexcelファイルのカラム名がTrueだったら
                        if select_item_list == selection_excel_file_columnslist:
                            # 抽出項目の値をdict化
                            select_value_dicts[select_item_list] = selection_excel_file.loc[selection_excel_file_len, select_item_list]

                # 転記先のデータフレームにdict化した値を追記する
                posting_new_df = posting_new_df.append(
                    select_value_dicts, ignore_index=True)

        print("車種の列のセル：　" + str(contents_cell))
    #    print("転記先のデータフレームに車種のセルを追記: " + posting_cell)
        print(posting_new_df)

        # Excelファイル操作-----------------------------------------------------------------------------------------------------
        # 編集するExcelファイルのファイル名のみを取得
        input_original_excel_file_name = os.path.splitext(
            os.path.basename(input_original_excel_file_path))[0]+"_excerpt"

        input_original_excel_file_name_xlsx = input_original_excel_file_name+".xlsx"

        duplication_path_test = "./"

        posting_new_df.to_excel(
            duplication_path_test+input_original_excel_file_name_xlsx, index=False)

        # excelの列幅を調整する
        excel_width_just_file = openpyxl.load_workbook(
            filename=input_original_excel_file_name_xlsx)
        excel_width_just_sheet = excel_width_just_file.worksheets[0]

        for col in excel_width_just_sheet.columns:
            max_length = 0
            column = col[0].column

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2) * 1.2
            excel_width_just_sheet.column_dimensions[col[0]
                                                     .column_letter].width = adjusted_width

        excel_width_just_file.save(input_original_excel_file_name_xlsx)

    else:
        # メインウィンドウの作成
        root = tkinter.Tk()
        # 最前面に表示
        root.attributes('-topmost', True)
        # メインウィンドウを非表示
        root.withdraw()
        # ダイアログの表示
        messagebox.showerror('実行エラー', after_ecxel_path +
                             '\nは項目抽出できるExcelファイルではありません！')


# メインの連番自動付与処理---------------------------------------------
# フォルダ内Excelファイル数のカウント
count = 0
# 読み込むフォルダパスの指定 ./は相対パス(相対パス ＝＞ 現在位置からの相対的な位置関係を記述する方式)
path = "./"
# フォルダ内のファイル一覧取得
files = os.listdir(path)
# フォルダ内の拡張子が.xlsxだったらファイル名取得し、function実行
for file in files:
    # os.path.splitext(ファイル名)で拡張子とそれ以外に分割されたタプルが戻り値をして戻る
    file_extension = os.path.splitext(file)
    # タプルの[1]＝＞拡張子を指定し、.xlsxだったら関数実行
    if file_extension[1] == ".xlsx":
        serial_number_assignment(file)
        count += 1
# フォルダ内のExcelファイルが無かった時の処理
if count == 0:
    # メインウィンドウの作成
    root = tkinter.Tk()
    # 最前面に表示
    root.attributes('-topmost', True)
    # メインウィンドウを非表示
    root.withdraw()
    # ダイアログの表示
    messagebox.showerror('実行エラー', 'Excelファイルがフォルダ内にありません！')
    sys.exit(1)

# 自動連番付与された項目を抽出する処理-----------------------------------
# 読み込むフォルダパスの指定
after_path = "./"
# フォルダ内のファイル一覧取得
after_files = os.listdir(after_path)
# フォルダ内の拡張子が.xlsxだったらファイル名取得し、function実行
for after_file in after_files:
    file_extension = os.path.splitext(after_file)
    if file_extension[1] == ".xlsx":
        serial_number_selection(after_file)

# 処理終了ダイアログ処理-------------------------------------------------

# メインウィンドウの作成
root = tkinter.Tk()
# 最前面に表示
root.attributes('-topmost', True)
# メインウィンドウを非表示
root.withdraw()
# ダイアログの表示
messagebox.showinfo('実行完了ダイアログ', '連番自動付与処理が完了しました！')
